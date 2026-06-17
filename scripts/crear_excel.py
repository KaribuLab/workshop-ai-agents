#!/usr/bin/env python3
"""CLI del workshop: genera el Excel de validación de contratos a partir del cruce.

El agente arma el cruce (PDF vs CRM) como una lista JSON y este CLI lo formatea en .xlsx.

Uso:
    python scripts/crear_excel.py --input cruce.json
    cat cruce.json | python scripts/crear_excel.py

Por defecto el .xlsx se genera en el Escritorio (~/Desktop/resultados.xlsx). Se puede
cambiar con --output.

Cada fila del JSON admite estas claves (todas opcionales salvo lo que quieras mostrar):
    archivo, cliente, cliente_existe, estado_cliente, poliza_pdf, poliza_bdd,
    coincide_poliza, monto_pdf, monto_crm, observacion, estado

Si "estado" no viene, se deriva con la precedencia de las reglas del workshop.
"""
import argparse
import json
import os
import sys
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# (clave en el JSON, encabezado en el Excel, ancho, ¿es monto?)
COLUMNAS = [
    ("archivo",         "Archivo",          16, False),
    ("cliente",         "Cliente",          26, False),
    ("cliente_existe",  "Cliente existe",   14, False),
    ("estado_cliente",  "Estado cliente",   14, False),
    ("poliza_pdf",      "Póliza PDF",        12, False),
    ("poliza_bdd",      "Póliza BDD",        12, False),
    ("coincide_poliza", "Coincide póliza",  15, False),
    ("monto_pdf",       "Monto PDF",        16, True),
    ("monto_crm",       "Monto CRM",        16, True),
    ("estado",          "Estado",           12, False),
    ("observacion",     "Observación",      40, False),
]

NAVY = "172A4D"
COLOR_ESTADO = {
    "OK":      "C6EFCE",  # verde
    "REVISAR": "FFEB9C",  # ámbar
    "ERROR":   "FFC7CE",  # rojo
}
TEXTO_ESTADO = {
    "OK":      "1B7A3D",
    "REVISAR": "9C6500",
    "ERROR":   "9C0006",
}


def normaliza(valor):
    """Mayúsculas sin acentos: 'Revisar' -> 'REVISAR', 'Sí' -> 'SI'."""
    if valor is None:
        return ""
    s = unicodedata.normalize("NFKD", str(valor))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().upper()


def es_no(valor):
    return normaliza(valor) in ("NO", "FALSE", "0", "N")


def a_numero(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return valor
    s = str(valor).replace("$", "").replace(".", "").replace(",", "").strip()
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return None


def derivar_estado(fila):
    """Aplica la precedencia de reglas si la fila no trae 'estado'."""
    if es_no(fila.get("cliente_existe")):
        return "ERROR"
    if es_no(fila.get("coincide_poliza")):
        return "ERROR"
    mp, mc = a_numero(fila.get("monto_pdf")), a_numero(fila.get("monto_crm"))
    if mp is not None and mc is not None and mp != mc:
        return "REVISAR"
    if normaliza(fila.get("estado_cliente")) == "INACTIVO":
        return "REVISAR"
    return "OK"


def cargar_filas(args):
    if args.input and args.input != "-":
        with open(args.input, encoding="utf-8") as fh:
            data = json.load(fh)
    else:
        data = json.load(sys.stdin)
    # Acepta una lista, o un objeto con la lista bajo "contratos"/"filas"/"resultados".
    if isinstance(data, dict):
        for k in ("contratos", "filas", "resultados", "rows"):
            if isinstance(data.get(k), list):
                data = data[k]
                break
    if not isinstance(data, list):
        raise ValueError("El JSON debe ser una lista de filas (o un objeto que la contenga).")
    return data


def construir(filas, salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Validación"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor=NAVY)
    head_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Encabezado
    for ci, (_, titulo, ancho, _) in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=ci, value=titulo)
        cell.fill, cell.font, cell.border, cell.alignment = head_fill, head_font, border, center
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    contador = {"OK": 0, "REVISAR": 0, "ERROR": 0}

    for ri, fila in enumerate(filas, start=2):
        estado = normaliza(fila.get("estado")) or derivar_estado(fila)
        if estado not in contador:
            estado = "REVISAR"  # fallback defensivo
        contador[estado] += 1

        for ci, (clave, _titulo, _ancho, es_monto) in enumerate(COLUMNAS, start=1):
            if clave == "estado":
                valor = estado
            elif es_monto:
                valor = a_numero(fila.get(clave))
            else:
                valor = fila.get(clave, "")
            cell = ws.cell(row=ri, column=ci, value=valor)
            cell.border = border
            if es_monto:
                cell.number_format = '#,##0'
                cell.alignment = center
            elif clave == "observacion":
                cell.alignment = left
            else:
                cell.alignment = center
            if clave == "estado":
                cell.fill = PatternFill("solid", fgColor=COLOR_ESTADO[estado])
                cell.font = Font(bold=True, color=TEXTO_ESTADO[estado])

    # Fila de totales
    total_row = len(filas) + 2
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
    resumen = f"OK={contador['OK']}   REVISAR={contador['REVISAR']}   ERROR={contador['ERROR']}"
    estado_col = next(i for i, c in enumerate(COLUMNAS, start=1) if c[0] == "observacion")
    tcell = ws.cell(row=total_row, column=estado_col, value=resumen)
    tcell.font = Font(bold=True)
    tcell.alignment = Alignment(horizontal="left", vertical="center")

    # Encabezado congelado + autofiltro sobre las filas de datos
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{len(filas) + 1}"

    wb.save(salida)
    return contador


def salida_por_defecto():
    """Por defecto el Excel se genera en el Escritorio si existe; si no, en el directorio actual."""
    escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
    if os.path.isdir(escritorio):
        return os.path.join(escritorio, "resultados.xlsx")
    return "resultados.xlsx"


def main():
    ap = argparse.ArgumentParser(description="Genera el Excel de validación de contratos.")
    ap.add_argument("-i", "--input", help="JSON de entrada (default: stdin).")
    ap.add_argument("-o", "--output", default=None,
                    help="Archivo .xlsx de salida (default: ~/Desktop/resultados.xlsx).")
    args = ap.parse_args()

    salida = args.output or salida_por_defecto()

    try:
        filas = cargar_filas(args)
    except (OSError, ValueError, json.JSONDecodeError) as e:
        print(f"Error leyendo el JSON de entrada: {e}", file=sys.stderr)
        return 1

    if not filas:
        print("Advertencia: la lista de filas está vacía; se generará un Excel solo con encabezado.", file=sys.stderr)

    try:
        contador = construir(filas, salida)
    except Exception as e:  # noqa: BLE001
        print(f"Error generando el Excel: {e}", file=sys.stderr)
        return 1

    print(f"OK: {salida} generado con {len(filas)} fila(s)  "
          f"(OK={contador['OK']}  REVISAR={contador['REVISAR']}  ERROR={contador['ERROR']}).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
